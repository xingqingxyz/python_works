from openpyxl import Workbook,load_workbook
import datetime

wb = Workbook()
ws = wb.active
ws['A1'] = 42
ws.append([1,2,3])
ws['A2'] = datetime.datetime.now()
wb.save('E:/Desktop/sample.xlsx')

wb = load_workbook('E:/Desktop/sample.xlsx')
print(wb.sheetnames)
